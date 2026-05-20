from pathlib import Path

from app.engines.bp_engine import forecast_periods


def build_business_plan_workbook(project: dict, financials: dict, output_path: Path) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        _write_csv_fallback(project, financials, output_path)
        return

    wb = Workbook()

    # --- Cover sheet ---
    ws_cover = wb.active
    ws_cover.title = "Cover"
    ws_cover["A1"] = project.get("company_name", "Company")
    ws_cover["A1"].font = Font(bold=True, size=16)
    ws_cover["A2"] = f"Project type: {project.get('project_type', '')}"
    ws_cover["A3"] = f"Currency: {project.get('currency', 'EUR')}"
    ws_cover["A4"] = f"Fiscal year end: {project.get('fiscal_year_end', '')}"

    historical_periods = financials.get("periods", [])
    forecast = forecast_periods(historical_periods[-1]) if historical_periods else forecast_periods("FY2024")
    all_periods = historical_periods + forecast

    # --- Income Statement sheet ---
    ws_is = wb.create_sheet("Income Statement")
    _write_section(ws_is, "Income Statement", financials.get("income_statement", []), all_periods, historical_periods)

    # --- Balance Sheet sheet ---
    ws_bs = wb.create_sheet("Balance Sheet")
    _write_section(ws_bs, "Balance Sheet", financials.get("balance_sheet", []), all_periods, historical_periods)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))


def _write_section(ws, title: str, lines: list[dict], all_periods: list[str], historical_periods: list[str]) -> None:
    from openpyxl.styles import Font, PatternFill, Alignment

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    hist_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    fore_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=13)

    header_row = 3
    ws.cell(row=header_row, column=1, value="Line Item").font = Font(bold=True, color="FFFFFF")
    ws.cell(row=header_row, column=1).fill = header_fill

    for col_idx, period in enumerate(all_periods, start=2):
        cell = ws.cell(row=header_row, column=col_idx, value=period)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, line in enumerate(lines, start=header_row + 1):
        ws.cell(row=row_idx, column=1, value=line.get("name", ""))
        values = line.get("values", {})
        for col_idx, period in enumerate(all_periods, start=2):
            cell = ws.cell(row=row_idx, column=col_idx, value=values.get(period, 0.0))
            cell.fill = hist_fill if period in historical_periods else fore_fill
            cell.number_format = "#,##0"


def _write_csv_fallback(project: dict, financials: dict, output_path: Path) -> None:
    import csv

    csv_path = output_path.with_suffix(".csv")
    csv_path.parent.mkdir(parents=True, exist_ok=True)
    periods = financials.get("periods", [])
    rows = []
    for section_key in ("income_statement", "balance_sheet"):
        for line in financials.get(section_key, []):
            row = {"Line Item": line.get("name", "")}
            row.update(line.get("values", {}))
            rows.append(row)

    if rows:
        fieldnames = ["Line Item"] + periods
        with csv_path.open("w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
            writer.writeheader()
            writer.writerows(rows)
