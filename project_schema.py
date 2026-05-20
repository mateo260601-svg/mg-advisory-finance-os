import csv
import json
from pathlib import Path

from app.engines.historical_accounts_engine import build_basic_historical_pack
from app.services.extraction_service import list_project_documents
from app.services.project_service import project_dir


def normalize_project_financials(project_id: str) -> dict:
    documents = list_project_documents(project_id)
    extracted_rows = []

    for document in documents:
        if document.suffix.lower() == ".csv":
            extracted_rows.extend(_read_csv_rows(document))
        elif document.suffix.lower() in {".xlsx", ".xlsm"}:
            extracted_rows.extend(_read_xlsx_rows(document))

    normalized = build_basic_historical_pack(extracted_rows, [path.name for path in documents])
    output_path = project_dir(project_id) / "normalized" / "financials.json"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(normalized, indent=2), encoding="utf-8")
    return normalized


def load_normalized_financials(project_id: str) -> dict:
    output_path = project_dir(project_id) / "normalized" / "financials.json"
    if not output_path.exists():
        return normalize_project_financials(project_id)
    try:
        return json.loads(output_path.read_text(encoding="utf-8"))
    except Exception:
        return normalize_project_financials(project_id)


def _read_csv_rows(path: Path) -> list[dict]:
    rows = []
    try:
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            for row in reader:
                rows.append({str(key or "").strip(): value for key, value in row.items()})
    except Exception:
        return []
    return rows


def _read_xlsx_rows(path: Path) -> list[dict]:
    try:
        from openpyxl import load_workbook
    except Exception:
        return []

    rows = []
    try:
        workbook = load_workbook(path, data_only=True, read_only=True)
        sheet = workbook.worksheets[0]
        values = list(sheet.iter_rows(values_only=True))
        if not values:
            return []
        headers = [str(cell or "").strip() for cell in values[0]]
        for values_row in values[1:]:
            row = {}
            for index, header in enumerate(headers):
                if header:
                    row[header] = values_row[index] if index < len(values_row) else None
            rows.append(row)
    except Exception:
        return []
    return rows

